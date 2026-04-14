import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# ===== 样式定义 =====
def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

header_fill = PatternFill('solid', fgColor='1E88E5')      # 蓝
night_fill  = PatternFill('solid', fgColor='5E35B1')      # 紫（晚上）
food_fill   = PatternFill('solid', fgColor='FF7043')      # 橙（美食）
scenic_fill = PatternFill('solid', fgColor='43A047')      # 绿（景点）
trans_fill  = PatternFill('solid', fgColor='FB8C00')      # 橙黄（交通）
hotel_fill  = PatternFill('solid', fgColor='8E24AA')      # 紫（住宿）
alt_fill    = PatternFill('solid', fgColor='E3F2FD')

white_bold = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
black_bold = Font(name='微软雅黑', bold=True, color='333333', size=10)
black_norm = Font(name='微软雅黑', color='333333', size=10)
red_bold   = Font(name='微软雅黑', bold=True, color='E53935', size=11)
title_font = Font(name='微软雅黑', bold=True, color='1565C0', size=18)
sub_font   = Font(name='微软雅黑', bold=True, color='1565C0', size=13)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== Sheet 1: 行程总览 =====
ws1 = wb.active
ws1.title = '30小时行程总览'
ws1.sheet_view.showGridLines = False

# 列宽
for col in ['A','B','C','D','E','F']:
    ws1.column_dimensions[col].width = 18
ws1.column_dimensions['D'].width = 35
ws1.column_dimensions['E'].width = 25

# 标题
ws1.merge_cells('A1:F1')
ws1['A1'] = '🌊 厦门30小时精华攻略 🌊'
ws1['A1'].font = title_font
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:F2')
ws1['A2'] = '到达：厦门北站 ｜ 游玩时间：30小时'
ws1['A2'].font = Font(name='微软雅黑', color='666666', size=11)
ws1['A2'].alignment = center
ws1.row_dimensions[2].height = 25

# 表头
headers = ['时间', '时段', '类型', '行程安排', '地点/交通', '备注']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = white_bold; c.fill = header_fill
    c.alignment = center; c.border = thin_border()
ws1.row_dimensions[4].height = 28

# 行程数据
schedule = [
    # Day 1 - 晚上到达
    ('18:00-19:00', '抵达', '交通', '厦门北站 → 酒店/民宿', '地铁1号线→镇海路站', '约50分钟，票价¥7'),
    ('19:00-20:00', '晚上', '住宿', '办理入住，稍作休息', '中山路/曾厝垵附近', '推荐住中山路，出行方便'),
    ('20:00-22:00', '晚上', '美食', '中山路步行街逛吃', '中山路', '沙茶面、土笋冻、花生汤'),
    ('22:00-23:00', '晚上', '夜景', '鹭江夜游或海边散步', '鹭江道/轮渡码头', '看鼓浪屿夜景'),
    
    # Day 2 - 上午
    ('07:30-08:30', '上午', '美食', '早餐：沙茶面/面线糊', '八市或酒店附近', '当地人推荐的古早味'),
    ('08:30-12:00', '上午', '景点', '鼓浪屿半日游', '鼓浪屿', '船票¥35，建议买早班船'),
    ('08:30-09:00', '上午', '交通', '轮渡码头→鼓浪屿', '东渡邮轮中心', '提前在"厦门轮渡"公众号购票'),
    ('09:00-12:00', '上午', '景点', '日光岩+菽庄花园+钢琴博物馆', '鼓浪屿', '日光岩¥50，菽庄花园¥30'),
    
    # Day 2 - 中午
    ('12:00-13:30', '中午', '美食', '鼓浪屿午餐', '龙头路', '海蛎煎、鱼丸汤、烧肉粽'),
    ('13:30-14:30', '中午', '景点', '漫步鼓浪屿小巷', '鼓浪屿', '最美转角、万国建筑'),
    ('14:30-15:00', '中午', '交通', '返回厦门岛', '轮渡', '三丘田码头或内厝澳码头'),
    
    # Day 2 - 下午
    ('15:00-17:30', '下午', '景点', '南普陀寺+厦门大学', '思明区', '南普陀免费，厦大需预约'),
    ('15:00-16:00', '下午', '景点', '南普陀寺', '思明南路', '素饼很好吃，可买伴手礼'),
    ('16:00-17:30', '下午', '景点', '厦门大学（外观/入内）', '思明南路', '需提前3天在U厦大预约'),
    ('17:30-18:00', '下午', '交通', '前往白城沙滩', '公交/打车', '看日落的最佳地点'),
    ('18:00-19:00', '傍晚', '景点', '白城沙滩日落', '环岛路', '免费，拍照绝美'),
    
    # Day 2 - 晚上
    ('19:00-20:30', '晚上', '美食', '曾厝垵晚餐', '曾厝垵', '文艺渔村，小吃很多'),
    ('20:30-22:00', '晚上', '夜景', '环岛路骑行/散步', '环岛路', '吹海风，看夜景'),
    
    # Day 3 - 上午（最后一天）
    ('08:00-09:00', '上午', '美食', '早餐：花生汤+油条', '思北花生汤', '老字号，人均¥15'),
    ('09:00-11:30', '上午', '景点', '植物园（雨林世界+多肉区）', '虎园路', '门票¥30，网红打卡地'),
    ('11:30-12:30', '中午', '美食', '午餐：海鲜大排档', '小眼镜大排档/202', '新鲜实惠，人均¥80'),
    ('12:30-13:30', '中午', '购物', '买伴手礼', '中山路/八市', '馅饼、鱼干、茶叶'),
    ('13:30-14:30', '下午', '交通', '前往厦门北站', '地铁1号线', '预留充足时间'),
    ('15:00', '结束', '返程', '离开厦门', '厦门北站', '结束愉快旅程！'),
]

row = 5
for time_slot, period, type_, activity, location, note in schedule:
    ws1.row_dimensions[row].height = 35
    
    # 类型颜色
    type_fill = {'交通': trans_fill, '住宿': hotel_fill, '美食': food_fill, 
                 '景点': scenic_fill, '晚上': night_fill, '傍晚': night_fill}.get(type_, alt_fill)
    
    ws1.cell(row=row, column=1, value=time_slot).font = black_bold
    ws1.cell(row=row, column=1).alignment = center
    ws1.cell(row=row, column=1).border = thin_border()
    
    ws1.cell(row=row, column=2, value=period).font = black_norm
    ws1.cell(row=row, column=2).alignment = center
    ws1.cell(row=row, column=2).border = thin_border()
    
    ws1.cell(row=row, column=3, value=type_).font = black_bold
    ws1.cell(row=row, column=3).fill = type_fill
    ws1.cell(row=row, column=3).alignment = center
    ws1.cell(row=row, column=3).border = thin_border()
    
    ws1.cell(row=row, column=4, value=activity).font = black_norm
    ws1.cell(row=row, column=4).alignment = left
    ws1.cell(row=row, column=4).border = thin_border()
    
    ws1.cell(row=row, column=5, value=location).font = black_bold
    ws1.cell(row=row, column=5).alignment = left
    ws1.cell(row=row, column=5).border = thin_border()
    
    ws1.cell(row=row, column=6, value=note).font = Font(name='微软雅黑', color='666666', size=9)
    ws1.cell(row=row, column=6).alignment = left
    ws1.cell(row=row, column=6).border = thin_border()
    
    row += 1

# ===== Sheet 2: 交通指南 =====
ws2 = wb.create_sheet('交通指南')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 20

ws2.merge_cells('A1:D1')
ws2['A1'] = '🚇 厦门北站交通指南'
ws2['A1'].font = title_font
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 38

# 北站到市区
ws2.merge_cells('A3:D3')
ws2['A3'] = '厦门北站 → 市区'
ws2['A3'].font = sub_font
ws2['A3'].alignment = left
ws2.row_dimensions[3].height = 25

transport = [
    ('地铁1号线', '最快推荐', '厦门北站→镇海路站（中山路）', '约50分钟，¥7'),
    ('BRT快1线', '经济实惠', '厦门北站→第一码头站', '约60分钟，¥4'),
    ('出租车/网约车', '舒适直达', '直达酒店/民宿', '约¥80-120，40分钟'),
    ('机场大巴', '备选方案', '厦门北站→轮渡/中山路', '约60分钟，¥15'),
]

for col, h in enumerate(['交通方式', '推荐度', '路线', '费用/时间'], 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font = white_bold; c.fill = trans_fill
    c.alignment = center; c.border = thin_border()
ws2.row_dimensions[4].height = 26

for i, (mode, rec, route, cost) in enumerate(transport, 5):
    ws2.row_dimensions[i].height = 28
    for col, val in enumerate([mode, rec, route, cost], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = black_norm if col != 1 else black_bold
        c.alignment = center if col in [1,2,4] else left
        c.border = thin_border()

# 市内交通
ws2.merge_cells('A10:D10')
ws2['A10'] = '市内交通'
ws2['A10'].font = sub_font
ws2['A10'].alignment = left
ws2.row_dimensions[10].height = 25

for col, h in enumerate(['交通方式', '适用场景', '说明', '费用'], 1):
    c = ws2.cell(row=11, column=col, value=h)
    c.font = white_bold; c.fill = trans_fill
    c.alignment = center; c.border = thin_border()
ws2.row_dimensions[11].height = 26

city_trans = [
    ('地铁', '长途跨区', '1/2/3号线，覆盖主要景点', '¥2-7'),
    ('公交', '经济实惠', '线路多，部分景点直达', '¥1-2'),
    ('BRT', '快速通行', '高架专用道，不堵车', '¥0.5-4'),
    ('出租车', '舒适便捷', '起步价¥10，适合短途', '按里程计费'),
    ('共享单车', '短途/环岛', '环岛路骑行首选', '¥1.5/半小时'),
    ('轮渡', '去鼓浪屿', '东渡邮轮中心→鼓浪屿', '¥35往返'),
]

for i, (mode, scene, desc, cost) in enumerate(city_trans, 12):
    ws2.row_dimensions[i].height = 26
    for col, val in enumerate([mode, scene, desc, cost], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = black_norm if col != 1 else black_bold
        c.alignment = center if col in [1,2,4] else left
        c.border = thin_border()

# ===== Sheet 3: 景点门票 =====
ws3 = wb.create_sheet('景点门票')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 30

ws3.merge_cells('A1:D1')
ws3['A1'] = '🎫 景点门票信息'
ws3['A1'].font = title_font
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 38

for col, h in enumerate(['景点名称', '门票', '预约方式', '建议游玩'], 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = white_bold; c.fill = scenic_fill
    c.alignment = center; c.border = thin_border()
ws3.row_dimensions[3].height = 26

tickets = [
    ('鼓浪屿轮渡', '¥35', '厦门轮渡+ 公众号', '半天'),
    ('日光岩', '¥50', '现场或携程/美团', '1小时'),
    ('菽庄花园', '¥30', '现场或携程/美团', '1小时'),
    ('南普陀寺', '免费', '南普陀寺公众号预约', '1小时'),
    ('厦门大学', '免费', 'U厦大 公众号预约', '1.5小时'),
    ('植物园', '¥30', '厦门植物园公众号', '2-3小时'),
    ('环岛路', '免费', '无需预约', '2小时'),
    ('曾厝垵', '免费', '无需预约', '2小时'),
    ('中山路', '免费', '无需预约', '2小时'),
]

for i, (name, price, book, time) in enumerate(tickets, 4):
    ws3.row_dimensions[i].height = 26
    for col, val in enumerate([name, price, book, time], 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = black_norm if col != 1 else black_bold
        c.alignment = center if col in [2,4] else left
        c.border = thin_border()

# ===== Sheet 4: 美食推荐 =====
ws4 = wb.create_sheet('美食推荐')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 15

ws4.merge_cells('A1:D1')
ws4['A1'] = '🍜 厦门必吃美食'
ws4['A1'].font = title_font
ws4['A1'].alignment = center
ws4.row_dimensions[1].height = 38

for col, h in enumerate(['美食名称', '推荐店铺', '特色说明', '人均'], 1):
    c = ws4.cell(row=3, column=col, value=h)
    c.font = white_bold; c.fill = food_fill
    c.alignment = center; c.border = thin_border()
ws4.row_dimensions[3].height = 26

foods = [
    ('沙茶面', '四里沙茶面/乌糖沙茶面', '厦门招牌，浓郁沙茶汤底', '¥15-30'),
    ('海蛎煎', '康家龙头/莲欢海蛎煎', '外酥里嫩，海鲜鲜美', '¥20-35'),
    ('土笋冻', '天河西门土笋冻', '特色闽南小吃，Q弹爽口', '¥15-25'),
    ('花生汤', '思北花生汤/黄则和', '甜而不腻，配油条绝了', '¥10-15'),
    ('面线糊', '浮屿面线糊', '早餐首选，暖胃暖心', '¥10-20'),
    ('烧肉粽', '1980烧肉粽', '料多实在，咸香可口', '¥15-25'),
    ('鱼丸汤', '原巷口鱼丸', '手工鱼丸，汤鲜味美', '¥15-25'),
    ('姜母鸭', '好德来姜母鸭', '滋补暖身，香气四溢', '¥80-120'),
    ('海鲜大排档', '小眼镜/202大排档', '新鲜实惠，品种丰富', '¥80-150'),
    ('馅饼', '南普陀素饼/汪记', '伴手礼首选，酥香甜', '¥20-40'),
]

for i, (name, shop, desc, price) in enumerate(foods, 4):
    ws4.row_dimensions[i].height = 28
    for col, val in enumerate([name, shop, desc, price], 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.font = black_norm if col != 1 else black_bold
        c.alignment = center if col == 4 else left
        c.border = thin_border()

# ===== Sheet 5: 费用预算 =====
ws5 = wb.create_sheet('费用预算')
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 15
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 15

ws5.merge_cells('A1:D1')
ws5['A1'] = '💰 30小时费用预算'
ws5['A1'].font = title_font
ws5['A1'].alignment = center
ws5.row_dimensions[1].height = 38

for col, h in enumerate(['项目', '经济型', '舒适型', '豪华型'], 1):
    c = ws5.cell(row=3, column=col, value=h)
    c.font = white_bold; c.fill = PatternFill('solid', fgColor='43A047')
    c.alignment = center; c.border = thin_border()
ws5.row_dimensions[3].height = 26

budget = [
    ('住宿（1晚）', '¥150-250', '¥300-500', '¥600-1200'),
    ('交通', '¥50-80', '¥100-150', '¥200-300'),
    ('门票', '¥150', '¥150', '¥200'),
    ('餐饮', '¥150-200', '¥300-400', '¥500-800'),
    ('伴手礼', '¥100', '¥200', '¥400'),
    ('合计', '¥600-880', '¥1050-1400', '¥1900-2900'),
]

for i, (item, eco, comf, lux) in enumerate(budget, 4):
    ws5.row_dimensions[i].height = 28
    is_total = item == '合计'
    for col, val in enumerate([item, eco, comf, lux], 1):
        c = ws5.cell(row=i, column=col, value=val)
        c.font = Font(name='微软雅黑', bold=True, color='E53935' if is_total else '333333', size=11 if is_total else 10)
        c.alignment = center
        c.border = thin_border()
        if is_total:
            c.fill = PatternFill('solid', fgColor='FFEBEE')

# ===== Sheet 6: 注意事项 =====
ws6 = wb.create_sheet('注意事项')
ws6.sheet_view.showGridLines = False
ws6.column_dimensions['A'].width = 80

ws6.merge_cells('A1:A1')
ws6['A1'] = '⚠️ 厦门旅游注意事项'
ws6['A1'].font = title_font
ws6['A1'].alignment = center
ws6.row_dimensions[1].height = 38

tips = [
    '',
    '📱 提前预约：',
    '   • 鼓浪屿船票：提前在"厦门轮渡+"公众号购买，建议买早班船（8:10或8:30）',
    '   • 厦门大学：提前3天在"U厦大"公众号预约，名额有限',
    '   • 南普陀寺：提前1天在公众号预约',
    '',
    '🌤️ 天气穿衣：',
    '   • 厦门3月气温约15-22℃，早晚凉，建议带薄外套',
    '   • 海边风大，注意防风',
    '   • 记得带防晒霜和墨镜',
    '',
    '🚫 避坑指南：',
    '   • 鼓浪屿上不要买所谓的"特产"，价格虚高',
    '   • 曾厝垵小吃价格偏高，适量尝试即可',
    '   • 海鲜大排档先问价再点菜，避免被宰',
    '   • 不要轻信路边拉客的"一日游"',
    '',
    '💡 实用Tips：',
    '   • 地铁/公交可用支付宝乘车码',
    '   • 鼓浪屿建议穿舒适的鞋子，需要走很多路',
    '   • 环岛路骑行推荐傍晚，看日落超美',
    '   • 八市（第八市场）是本地人买菜的地方，小吃正宗又便宜',
    '   • 伴手礼推荐：南普陀素饼、日光岩馅饼、鱼干、茶叶',
    '',
    '📞 紧急联系：',
    '   • 旅游投诉：12345',
    '   • 急救：120',
    '   • 报警：110',
]

for i, tip in enumerate(tips, 3):
    ws6.row_dimensions[i].height = 22 if tip else 10
    c = ws6.cell(row=i, column=1, value=tip)
    c.font = black_norm if not tip.startswith('  ') else Font(name='微软雅黑', color='666666', size=10)
    c.alignment = left

out = '/Users/xiaoxiami/.qclaw/workspace/厦门30小时攻略.xlsx'
wb.save(out)
print(f"厦门攻略生成完成！")
