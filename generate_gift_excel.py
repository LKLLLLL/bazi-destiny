import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

wb = openpyxl.Workbook()

# ===== 数据 =====
gifts = [
    ('罗云慧', 600), ('罗发标', 600), ('王滨', 1000), ('罗建国', 1000),
    ('莲姨', 1000), ('罗静楚', 600), ('小叔小婶', 2000), ('三叔公', 2000),
    ('巧珊', 500), ('王萍', 1000), ('婕姐姐', 600), ('王东', 500),
    ('温小敏', 500), ('张文响', 600), ('王小洪', 1000), ('钟喜萍', 1000),
    ('朱兴伟', 300), ('王秋菊', 300), ('张计晴', 600), ('叶宝嫦', 300),
    ('周佛金', 200), ('叶房桂', 600), ('叶冰峰', 1000), ('王武', 600),
    ('张挚华', 2000), ('王超', 500), ('叶永坚', 600), ('张立浩', 600),
    ('女王群9人', 2700), ('水叔', 600), ('林贵凯', 369), ('谢萍', 200),
    ('张春菊', 300), ('张冬娣', 600), ('李丹琪', 200), ('吴胤', 600),
    ('张战', 600), ('没写名字的', 5000),
]

total = sum(a for _, a in gifts)
count = len(gifts)
avg = total / count
max_gift = max(gifts, key=lambda x: x[1])
min_gift = min(gifts, key=lambda x: x[1])

def tier(a):
    if a >= 2000: return '2000元以上'
    elif a >= 1000: return '1000-1999元'
    elif a >= 600: return '600-999元'
    elif a >= 300: return '300-599元'
    else: return '200-299元'

# ===== 样式 =====
def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

header_fill  = PatternFill('solid', fgColor='C0392B')
summary_fill = PatternFill('solid', fgColor='FDEBD0')
alt_fill     = PatternFill('solid', fgColor='FEF9F9')
tier_fills = {
    '2000元以上': PatternFill('solid', fgColor='FADBD8'),
    '1000-1999元': PatternFill('solid', fgColor='FDEBD0'),
    '600-999元':  PatternFill('solid', fgColor='D5F5E3'),
    '300-599元':  PatternFill('solid', fgColor='D6EAF8'),
    '200-299元':  PatternFill('solid', fgColor='EAECEE'),
}

white_bold = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
red_bold   = Font(name='微软雅黑', bold=True, color='C0392B', size=12)
black_bold = Font(name='微软雅黑', bold=True, color='333333', size=10)
black_norm = Font(name='微软雅黑', color='333333', size=10)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left', vertical='center')

# ===== Sheet 1: 礼金明细 =====
ws1 = wb.active
ws1.title = '礼金明细'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 16
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 28

# 大标题
ws1.merge_cells('A1:D1')
ws1['A1'] = '宝宝百日礼金明细'
ws1['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=16)
ws1['A1'].alignment = center

# 表头
headers = ['序号', '姓名', '礼金金额（元）', '金额档位']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = white_bold
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border()

# 数据行
for i, (name, amount) in enumerate(gifts, 1):
    row = i + 2
    ws1.row_dimensions[row].height = 22
    t = tier(amount)
    fill = tier_fills[t] if i % 2 == 0 else alt_fill

    ws1.cell(row=row, column=1, value=i).font = black_norm
    ws1.cell(row=row, column=1).alignment = center
    ws1.cell(row=row, column=1).border = thin_border()

    ws1.cell(row=row, column=2, value=name).font = black_bold
    ws1.cell(row=row, column=2).alignment = left
    ws1.cell(row=row, column=2).border = thin_border()

    ws1.cell(row=row, column=3, value=amount).font = Font(name='微软雅黑', bold=True, color='C0392B', size=11)
    ws1.cell(row=row, column=3).number_format = '¥#,##0'
    ws1.cell(row=row, column=3).alignment = center
    ws1.cell(row=row, column=3).border = thin_border()
    ws1.cell(row=row, column=3).fill = tier_fills[t]

    ws1.cell(row=row, column=4, value=t).font = black_norm
    ws1.cell(row=row, column=4).alignment = center
    ws1.cell(row=row, column=4).border = thin_border()
    ws1.cell(row=row, column=4).fill = tier_fills[t]

# 合计行
total_row = count + 3
ws1.row_dimensions[total_row].height = 26
ws1.merge_cells(f'A{total_row}:B{total_row}')
ws1[f'A{total_row}'] = '合计'
ws1[f'A{total_row}'].font = white_bold
ws1[f'A{total_row}'].fill = header_fill
ws1[f'A{total_row}'].alignment = center
ws1[f'A{total_row}'].border = thin_border()

ws1[f'C{total_row}'] = total
ws1[f'C{total_row}'].font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=12)
ws1[f'C{total_row}'].number_format = '¥#,##0'
ws1[f'C{total_row}'].fill = header_fill
ws1[f'C{total_row}'].alignment = center
ws1[f'C{total_row}'].border = thin_border()

ws1[f'D{total_row}'] = f'共 {count} 位亲友'
ws1[f'D{total_row}'].font = white_bold
ws1[f'D{total_row}'].fill = header_fill
ws1[f'D{total_row}'].alignment = center
ws1[f'D{total_row}'].border = thin_border()

# ===== Sheet 2: 汇总统计 =====
ws2 = wb.create_sheet('汇总统计')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16

# 标题
ws2.merge_cells('A1:E1')
ws2['A1'] = '宝宝百日礼金 · 汇总统计'
ws2['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=15)
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 36

# 关键指标
ws2.merge_cells('A3:E3')
ws2['A3'] = '关键指标'
ws2['A3'].font = white_bold
ws2['A3'].fill = header_fill
ws2['A3'].alignment = center
ws2.row_dimensions[3].height = 26

key_stats = [
    ('礼金总计', f'¥{total:,}'),
    ('亲友人数', f'{count} 位'),
    ('人均礼金', f'¥{avg:.0f}'),
    ('最高礼金', f'{max_gift[0]}  ¥{max_gift[1]:,}'),
    ('最低礼金', f'{min_gift[0]}  ¥{min_gift[1]:,}'),
    ('500元以下人数', f'{sum(1 for _,a in gifts if a < 500)} 位'),
    ('1000元及以上人数', f'{sum(1 for _,a in gifts if a >= 1000)} 位'),
    ('2000元及以上人数', f'{sum(1 for _,a in gifts if a >= 2000)} 位'),
]

for i, (label, val) in enumerate(key_stats):
    row = i + 4
    ws2.row_dimensions[row].height = 24
    ws2.cell(row=row, column=1, value=label).font = black_bold
    ws2.cell(row=row, column=1).fill = summary_fill
    ws2.cell(row=row, column=1).alignment = left
    ws2.cell(row=row, column=1).border = thin_border()
    ws2.cell(row=row, column=2, value=val).font = Font(name='微软雅黑', bold=True, color='C0392B', size=11)
    ws2.cell(row=row, column=2).fill = summary_fill
    ws2.cell(row=row, column=2).alignment = center
    ws2.cell(row=row, column=2).border = thin_border()

# 档位统计表
ws2.merge_cells('A14:E14')
ws2['A14'] = '金额档位统计'
ws2['A14'].font = white_bold
ws2['A14'].fill = header_fill
ws2['A14'].alignment = center
ws2.row_dimensions[14].height = 26

tier_headers = ['金额档位', '人数', '占比', '礼金小计', '占总额比']
for col, h in enumerate(tier_headers, 1):
    ws2.cell(row=15, column=col, value=h).font = white_bold
    ws2.cell(row=15, column=col).fill = PatternFill('solid', fgColor='E74C3C')
    ws2.cell(row=15, column=col).alignment = center
    ws2.cell(row=15, column=col).border = thin_border()
ws2.row_dimensions[15].height = 22

order = ['2000元以上', '1000-1999元', '600-999元', '300-599元', '200-299元']
tier_count_map = {o: 0 for o in order}
tier_sum_map   = {o: 0 for o in order}
for _, a in gifts:
    t = tier(a)
    tier_count_map[t] += 1
    tier_sum_map[t]   += a

for i, t in enumerate(order):
    row = 16 + i
    ws2.row_dimensions[row].height = 22
    tc = tier_count_map[t]
    ts = tier_sum_map[t]
    fill = tier_fills[t]

    ws2.cell(row=row, column=1, value=t).font = black_bold
    ws2.cell(row=row, column=1).fill = fill
    ws2.cell(row=row, column=1).alignment = center
    ws2.cell(row=row, column=1).border = thin_border()

    ws2.cell(row=row, column=2, value=tc).font = black_bold
    ws2.cell(row=row, column=2).fill = fill
    ws2.cell(row=row, column=2).alignment = center
    ws2.cell(row=row, column=2).border = thin_border()

    ws2.cell(row=row, column=3, value=f'{tc/count*100:.1f}%').font = black_norm
    ws2.cell(row=row, column=3).fill = fill
    ws2.cell(row=row, column=3).alignment = center
    ws2.cell(row=row, column=3).border = thin_border()

    ws2.cell(row=row, column=4, value=ts).font = Font(name='微软雅黑', bold=True, color='C0392B', size=10)
    ws2.cell(row=row, column=4).number_format = '¥#,##0'
    ws2.cell(row=row, column=4).fill = fill
    ws2.cell(row=row, column=4).alignment = center
    ws2.cell(row=row, column=4).border = thin_border()

    ws2.cell(row=row, column=5, value=f'{ts/total*100:.1f}%').font = black_norm
    ws2.cell(row=row, column=5).fill = fill
    ws2.cell(row=row, column=5).alignment = center
    ws2.cell(row=row, column=5).border = thin_border()

# 合计行
ws2.row_dimensions[21].height = 24
ws2.cell(row=21, column=1, value='合计').font = white_bold
ws2.cell(row=21, column=1).fill = header_fill
ws2.cell(row=21, column=1).alignment = center
ws2.cell(row=21, column=1).border = thin_border()
ws2.cell(row=21, column=2, value=count).font = white_bold
ws2.cell(row=21, column=2).fill = header_fill
ws2.cell(row=21, column=2).alignment = center
ws2.cell(row=21, column=2).border = thin_border()
ws2.cell(row=21, column=3, value='100%').font = white_bold
ws2.cell(row=21, column=3).fill = header_fill
ws2.cell(row=21, column=3).alignment = center
ws2.cell(row=21, column=3).border = thin_border()
ws2.cell(row=21, column=4, value=total).font = white_bold
ws2.cell(row=21, column=4).number_format = '¥#,##0'
ws2.cell(row=21, column=4).fill = header_fill
ws2.cell(row=21, column=4).alignment = center
ws2.cell(row=21, column=4).border = thin_border()
ws2.cell(row=21, column=5, value='100%').font = white_bold
ws2.cell(row=21, column=5).fill = header_fill
ws2.cell(row=21, column=5).alignment = center
ws2.cell(row=21, column=5).border = thin_border()

# ===== Sheet 3: TOP排行 =====
ws3 = wb.create_sheet('礼金排行')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 14

ws3.merge_cells('A1:D1')
ws3['A1'] = '礼金排行榜（从高到低）'
ws3['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=15)
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 36

rank_headers = ['排名', '姓名', '礼金金额（元）', '金额档位']
for col, h in enumerate(rank_headers, 1):
    ws3.cell(row=2, column=col, value=h).font = white_bold
    ws3.cell(row=2, column=col).fill = header_fill
    ws3.cell(row=2, column=col).alignment = center
    ws3.cell(row=2, column=col).border = thin_border()
ws3.row_dimensions[2].height = 26

sorted_gifts = sorted(gifts, key=lambda x: x[1], reverse=True)
rank_fill_top3 = [
    PatternFill('solid', fgColor='FFD700'),
    PatternFill('solid', fgColor='C0C0C0'),
    PatternFill('solid', fgColor='CD7F32'),
]

for i, (name, amount) in enumerate(sorted_gifts, 1):
    row = i + 2
    ws3.row_dimensions[row].height = 22
    t = tier(amount)
    fill = rank_fill_top3[i-1] if i <= 3 else tier_fills[t]
    font = Font(name='微软雅黑', bold=True, color='333333', size=10) if i <= 3 else black_norm

    medal = {1: '🥇', 2: '🥈', 3: '🥉'}.get(i, str(i))
    ws3.cell(row=row, column=1, value=medal if i <= 3 else i).font = font
    ws3.cell(row=row, column=1).fill = fill
    ws3.cell(row=row, column=1).alignment = center
    ws3.cell(row=row, column=1).border = thin_border()

    ws3.cell(row=row, column=2, value=name).font = font
    ws3.cell(row=row, column=2).fill = fill
    ws3.cell(row=row, column=2).alignment = left
    ws3.cell(row=row, column=2).border = thin_border()

    ws3.cell(row=row, column=3, value=amount).font = Font(name='微软雅黑', bold=True, color='C0392B', size=11)
    ws3.cell(row=row, column=3).number_format = '¥#,##0'
    ws3.cell(row=row, column=3).fill = fill
    ws3.cell(row=row, column=3).alignment = center
    ws3.cell(row=row, column=3).border = thin_border()

    ws3.cell(row=row, column=4, value=t).font = font
    ws3.cell(row=row, column=4).fill = fill
    ws3.cell(row=row, column=4).alignment = center
    ws3.cell(row=row, column=4).border = thin_border()

# 保存
out = '/Users/xiaoxiami/.qclaw/workspace/宝宝百日礼金表格.xlsx'
wb.save(out)
print(f"Excel生成完成！总礼金: ¥{total:,}，共{count}人，人均¥{avg:.0f}")
