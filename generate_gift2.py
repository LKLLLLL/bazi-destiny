import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ===== 礼金数据（手动整理）=====
gifts = [
    ('小叔', 6666),
    ('张小可夫妇', 600),
    ('叶向红', 1200),
    ('水叔', 2000),
    ('官志碧', 2400),
    ('叶彩萍浩叔', 2600),
    ('官罗勇苏友谊大姐（慧）', 2000),
    ('黄伟平罗素芬丹姑婆', 1000),
    ('李石娣张天海', 600),
    ('张飞丹', 1000),
    ('钟国成叶巧玲', 200),
    ('笑姨婆', 1000),
    ('欧仲平标叔', 1000),
    ('大伯', 2000),
    ('叶丹梦姑', 600),
    ('三叔', 1000),
    ('欧克念竻叔', 500),
    ('阿嬷', 1000),
    ('苏志军', 900),
    ('谭柏权', 200),
    ('娜姑', 1000),
    ('苏志勇三妹嫂', 200),
    ('大德', 600),
    ('苏志能季长浓', 600),
    ('张伟', 600),
    ('赖建生', 500),
    ('罗汉林', 200),
    ('黄雪芳', 600),
    ('黄炜伦', 1000),
    ('叶建平', 2000),
    ('刘伟苑', 600),
    ('莲姨', 1000),
    ('罗可知', 1000),
    ('欧六妹', 600),
    ('璇姨', 600),
    ('罗圳烈叶媚卿', 600),
    ('张秋凤', 500),
    ('罗捷思', 600),
    ('谭旭颖', 600),
    ('香港阿姑', 2000),
    ('罗方思二姑', 500),
    ('叶友艺', 1000),
    ('赖志聪', 1000),
    ('安老婆', 1000),
    ('杨敏', 600),
    ('叶继东', 600),
    ('谢风琴', 300),
    ('叶国立', 600),
    ('好伯', 600),
    ('根叔妈妈', 200),
    ('赖志忠', 200),
    ('邓书才', 2380),
    ('冰峰叔', 2000),
    ('策叔', 600),
    ('罗发奎叶展隆街海叔', 300),
    ('罗发坤季成阿霞婶(啊洲妈妈)', 300),
    ('罗发敏长娣婶', 200),
    ('梁谷花', 2000),
    ('罗竻耀雨峰老表×2', 1200),
    ('梁齐花', 600),
    ('罗丙贤张飞群', 300),
    ('梁静花', 600),
    ('罗东青小柔', 300),
    ('梁福花', 600),
    ('罗国房芳叔公', 1000),
    ('季伍英', 300),
    ('飘叔', 300),
    ('典叔', 1000),
    ('季广财', 200),
    ('糯米爷爷奶奶', 2000),
    ('何彩燕', 1000),
    ('谢荷岸', 200),
    ('糯米外公外婆', 13800),
    ('来哥', 666),
    ('叶映川', 1000),
    ('叶剑华', 600),
    ('黄丽英', 1500),
    ('张文力', 600),
    ('谢春球', 500),
    ('赖润霞张二叔', 600),
    ('叶健', 1000),
    ('黄玲张方犬', 1000),
    ('林贵铠', 1000),
    ('叶小红季广凯', 200),
    ('张梓凌', 1000),
    ('江丽蓉叶发兴', 600),
    ('桂秀娴', 300),
    ('李伟玲', 300),
    ('刘概', 600),
    ('徐慧和', 200),
    ('坚叔', 800),
]

total = sum(a for _, a in gifts)
count = len(gifts)
avg = total / count
max_gift = max(gifts, key=lambda x: x[1])
min_gift = min(gifts, key=lambda x: x[1])

def tier(a):
    if a >= 5000: return '5000元以上'
    elif a >= 2000: return '2000-4999元'
    elif a >= 1000: return '1000-1999元'
    elif a >= 600: return '600-999元'
    elif a >= 300: return '300-599元'
    else: return '200-299元'

# ===== 样式 =====
def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

header_fill = PatternFill('solid', fgColor='C0392B')
summary_fill = PatternFill('solid', fgColor='FDEBD0')
tier_fills = {
    '5000元以上':  PatternFill('solid', fgColor='F1948A'),
    '2000-4999元': PatternFill('solid', fgColor='FADBD8'),
    '1000-1999元': PatternFill('solid', fgColor='FDEBD0'),
    '600-999元':   PatternFill('solid', fgColor='D5F5E3'),
    '300-599元':   PatternFill('solid', fgColor='D6EAF8'),
    '200-299元':   PatternFill('solid', fgColor='EAECEE'),
}

white_bold = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
black_bold = Font(name='微软雅黑', bold=True, color='333333', size=10)
black_norm = Font(name='微软雅黑', color='333333', size=10)
red_bold   = Font(name='微软雅黑', bold=True, color='C0392B', size=11)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left', vertical='center')

# ===== Sheet 1: 礼金明细 =====
ws1 = wb.active
ws1.title = '礼金明细'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16

ws1.merge_cells('A1:D1')
ws1['A1'] = '宝宝百日礼金明细'
ws1['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=16)
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 38

for col, h in enumerate(['序号', '姓名', '礼金金额（元）', '金额档位'], 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = white_bold; c.fill = header_fill
    c.alignment = center; c.border = thin_border()
ws1.row_dimensions[2].height = 26

for i, (name, amount) in enumerate(gifts, 1):
    row = i + 2
    ws1.row_dimensions[row].height = 22
    t = tier(amount)
    f = tier_fills[t]

    ws1.cell(row=row, column=1, value=i).font = black_norm
    ws1.cell(row=row, column=1).alignment = center
    ws1.cell(row=row, column=1).border = thin_border()

    ws1.cell(row=row, column=2, value=name).font = black_bold
    ws1.cell(row=row, column=2).alignment = left
    ws1.cell(row=row, column=2).border = thin_border()

    ws1.cell(row=row, column=3, value=amount).font = red_bold
    ws1.cell(row=row, column=3).number_format = '¥#,##0'
    ws1.cell(row=row, column=3).alignment = center
    ws1.cell(row=row, column=3).fill = f
    ws1.cell(row=row, column=3).border = thin_border()

    ws1.cell(row=row, column=4, value=t).font = black_norm
    ws1.cell(row=row, column=4).alignment = center
    ws1.cell(row=row, column=4).fill = f
    ws1.cell(row=row, column=4).border = thin_border()

# 合计行
tr = count + 3
ws1.row_dimensions[tr].height = 28
ws1.merge_cells(f'A{tr}:B{tr}')
ws1[f'A{tr}'] = '合  计'
ws1[f'A{tr}'].font = white_bold; ws1[f'A{tr}'].fill = header_fill
ws1[f'A{tr}'].alignment = center; ws1[f'A{tr}'].border = thin_border()
ws1[f'C{tr}'] = total
ws1[f'C{tr}'].font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=12)
ws1[f'C{tr}'].number_format = '¥#,##0'
ws1[f'C{tr}'].fill = header_fill; ws1[f'C{tr}'].alignment = center; ws1[f'C{tr}'].border = thin_border()
ws1[f'D{tr}'] = f'共 {count} 位亲友'
ws1[f'D{tr}'].font = white_bold; ws1[f'D{tr}'].fill = header_fill
ws1[f'D{tr}'].alignment = center; ws1[f'D{tr}'].border = thin_border()

# ===== Sheet 2: 汇总统计 =====
ws2 = wb.create_sheet('汇总统计')
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16

ws2.merge_cells('A1:E1')
ws2['A1'] = '宝宝百日礼金 · 汇总统计'
ws2['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=15)
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 38

ws2.merge_cells('A3:E3')
ws2['A3'] = '关键指标'
ws2['A3'].font = white_bold; ws2['A3'].fill = header_fill
ws2['A3'].alignment = center; ws2.row_dimensions[3].height = 26

key_stats = [
    ('礼金总计', f'¥{total:,}'),
    ('亲友人数', f'{count} 位'),
    ('人均礼金', f'¥{avg:.0f}'),
    ('最高礼金', f'{max_gift[0]}  ¥{max_gift[1]:,}'),
    ('最低礼金', f'{min_gift[0]}  ¥{min_gift[1]:,}'),
    ('1000元及以上人数', f'{sum(1 for _,a in gifts if a >= 1000)} 位'),
    ('2000元及以上人数', f'{sum(1 for _,a in gifts if a >= 2000)} 位'),
    ('5000元及以上人数', f'{sum(1 for _,a in gifts if a >= 5000)} 位'),
]
for i, (label, val) in enumerate(key_stats):
    row = i + 4
    ws2.row_dimensions[row].height = 24
    ws2.cell(row=row, column=1, value=label).font = black_bold
    ws2.cell(row=row, column=1).fill = summary_fill
    ws2.cell(row=row, column=1).alignment = left
    ws2.cell(row=row, column=1).border = thin_border()
    ws2.cell(row=row, column=2, value=val).font = Font(name='微软雅黑', bold=True, color='C0392B', size=11)
    ws2.cell(row=row, column=2).fill = summary_fill
    ws2.cell(row=row, column=2).alignment = center
    ws2.cell(row=row, column=2).border = thin_border()

# 档位统计
ws2.merge_cells('A14:E14')
ws2['A14'] = '金额档位统计'
ws2['A14'].font = white_bold; ws2['A14'].fill = header_fill
ws2['A14'].alignment = center; ws2.row_dimensions[14].height = 26

for col, h in enumerate(['金额档位', '人数', '占比', '礼金小计', '占总额比'], 1):
    ws2.cell(row=15, column=col, value=h).font = white_bold
    ws2.cell(row=15, column=col).fill = PatternFill('solid', fgColor='E74C3C')
    ws2.cell(row=15, column=col).alignment = center
    ws2.cell(row=15, column=col).border = thin_border()
ws2.row_dimensions[15].height = 22

order = ['5000元以上', '2000-4999元', '1000-1999元', '600-999元', '300-599元', '200-299元']
tier_count_map = {o: 0 for o in order}
tier_sum_map   = {o: 0 for o in order}
for _, a in gifts:
    t = tier(a)
    tier_count_map[t] += 1
    tier_sum_map[t]   += a

for i, t in enumerate(order):
    row = 16 + i
    ws2.row_dimensions[row].height = 22
    tc = tier_count_map[t]; ts = tier_sum_map[t]
    f = tier_fills[t]
    for col, val in enumerate([t, tc, f'{tc/count*100:.1f}%', ts, f'{ts/total*100:.1f}%'], 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.fill = f; c.alignment = center; c.border = thin_border()
        if col == 4:
            c.number_format = '¥#,##0'
            c.font = Font(name='微软雅黑', bold=True, color='C0392B', size=10)
        else:
            c.font = black_bold if col == 1 else black_norm

# 合计行
ws2.row_dimensions[22].height = 26
for col, val in enumerate(['合计', count, '100%', total, '100%'], 1):
    c = ws2.cell(row=22, column=col, value=val)
    c.font = white_bold; c.fill = header_fill
    c.alignment = center; c.border = thin_border()
    if col == 4: c.number_format = '¥#,##0'

# ===== Sheet 3: 礼金排行 =====
ws3 = wb.create_sheet('礼金排行')
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16

ws3.merge_cells('A1:D1')
ws3['A1'] = '礼金排行榜（从高到低）'
ws3['A1'].font = Font(name='微软雅黑', bold=True, color='C0392B', size=15)
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 38

for col, h in enumerate(['排名', '姓名', '礼金金额（元）', '金额档位'], 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = white_bold; c.fill = header_fill
    c.alignment = center; c.border = thin_border()
ws3.row_dimensions[2].height = 26

top3_fills = [
    PatternFill('solid', fgColor='FFD700'),
    PatternFill('solid', fgColor='C0C0C0'),
    PatternFill('solid', fgColor='CD7F32'),
]

for i, (name, amount) in enumerate(sorted(gifts, key=lambda x: x[1], reverse=True), 1):
    row = i + 2
    ws3.row_dimensions[row].height = 22
    t = tier(amount)
    f = top3_fills[i-1] if i <= 3 else tier_fills[t]
    medal = {1:'🥇', 2:'🥈', 3:'🥉'}.get(i, i)

    ws3.cell(row=row, column=1, value=medal).font = black_bold
    ws3.cell(row=row, column=1).fill = f; ws3.cell(row=row, column=1).alignment = center; ws3.cell(row=row, column=1).border = thin_border()
    ws3.cell(row=row, column=2, value=name).font = black_bold
    ws3.cell(row=row, column=2).fill = f; ws3.cell(row=row, column=2).alignment = left; ws3.cell(row=row, column=2).border = thin_border()
    ws3.cell(row=row, column=3, value=amount).font = red_bold
    ws3.cell(row=row, column=3).number_format = '¥#,##0'
    ws3.cell(row=row, column=3).fill = f; ws3.cell(row=row, column=3).alignment = center; ws3.cell(row=row, column=3).border = thin_border()
    ws3.cell(row=row, column=4, value=t).font = black_norm
    ws3.cell(row=row, column=4).fill = f; ws3.cell(row=row, column=4).alignment = center; ws3.cell(row=row, column=4).border = thin_border()

out = '/Users/xiaoxiami/.qclaw/workspace/宝宝百日礼金表格2.xlsx'
wb.save(out)
print(f"完成！总礼金: ¥{total:,}，共{count}人，人均¥{avg:.0f}")
print(f"最高: {max_gift[0]} ¥{max_gift[1]:,}")
print(f"最低: {min_gift[0]} ¥{min_gift[1]:,}")
